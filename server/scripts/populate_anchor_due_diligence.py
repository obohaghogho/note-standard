import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

path = r'C:\Users\hp\Downloads\Updated Anchor Due Diligence document.xlsx'
out_path1 = r'C:\Users\hp\Downloads\Completed Anchor Questionnaire.xlsx'
out_path2 = path

wb = openpyxl.load_workbook(path, data_only=False)

ws1 = wb['Due Diligence for Anchor ']

def safe_set(r, col, val):
    cell = ws1.cell(r, col)
    if type(cell).__name__ == 'MergedCell':
        coord = cell.coordinate
        for rng in ws1.merged_cells.ranges:
            if coord in rng:
                ws1.cell(rng.min_row, rng.min_col).value = val
                return
    else:
        cell.value = val

def set_y(r):
    safe_set(r, 8, 'X')

def set_n(r):
    safe_set(r, 9, 'X')

# 1. COMPANY DETAILS
safe_set(3, 2, 'Jossy Digital Technologies Ltd')      # B3: Company Name
safe_set(3, 9, 'RC9586407')                            # I3: Company No
safe_set(4, 2, 'NoteStandard')                        # B4: Trading As
safe_set(5, 4, 'Nigeria (Incorporated). Operating in Nigeria, United Kingdom, United States, Ghana, Kenya') # D5
safe_set(6, 2, '2 June 2026')                          # B6: Date of Incorporation
safe_set(7, 2, '10 Winnie Okia Street, Effurun, Delta State, Nigeria') # B7: Registered Address
safe_set(8, 2, '330102')                              # B8: Post Code
safe_set(9, 2, '10 Winnie Okia Street, Effurun, Delta State, Nigeria') # B9: Correspondence Address
safe_set(10, 2, 'admin@notestandard.com')              # B10: Contact Email(s)
safe_set(11, 2, 'Attached (Certificate of Incorporation - RC9586407)') # B11
safe_set(12, 2, 'Attached (MEMAT / Company Constitution)')              # B12
safe_set(13, 2, 'Attached (Status Report / CAC Register detailing Director & Shareholders)') # B13
safe_set(14, 2, 'Financial Technology Platform (Operating through regulated financial infrastructure partners)') # B14

# 2. LEGAL ENTITY / OWNERSHIP
set_y(20) # Privately Owned
safe_set(21, 1, "(If Yes, provide details of shareholders or ultimate beneficial owners with a holding of 10% or more of the Entity's total shares composed of bearer shares): Aghogho Jossy Oboh (100% Ordinary Shares & UBO)")
safe_set(22, 1, 'Provide Legal Entity Identifier (LEI) if available: N/A (Private Limited Entity)')
safe_set(23, 1, 'Provide the full legal name of the ultimate parent: N/A (Jossy Digital Technologies Ltd is the ultimate entity)')
safe_set(24, 1, 'Jurisdiction of licensing authority: Nigeria (Corporate Affairs Commission & Central Bank of Nigeria framework)')
safe_set(25, 1, 'List the business areas applicable: Digital Payments, Multi-Currency Wallets, P2P Remittances, Treasury Platform, Ledger Engine')
set_y(27) # Non-resident customers >10%
safe_set(28, 1, 'If Yes, top 5 countries: United Kingdom, United States, Ghana, Kenya, European Union')
safe_set(29, 1, 'Number of employees: 1 Full-Time Founder & Managing Director (supported by specialized technical/compliance contractors)')
safe_set(30, 1, 'Total Assets: USD 250,000 (Pre-launch software, platform infrastructure & working capital)')
set_y(32) # Representative of all branches
safe_set(33, 1, 'N/A - Centralized corporate operations')
safe_set(34, 1, 'NoteStandard is an integrated real-time communications and financial technology platform currently in Private Testing, launching publicly on September 25, 2026.')

# 3. PRODUCTS / SERVICES
safe_set(37, 1, 'Multi-Currency Wallet, Digital Payments, Treasury Platform, Cross-Border Payments, Internal Ledger, Crypto Settlement Services through regulated partners, Merchant Payment Platform')
set_y(39) # Representative of all branches
safe_set(40, 1, 'N/A')
safe_set(41, 1, 'All financial services operate through tier 1-3 KYC verification and automated sanctions/PEP screening.')

# 4. GENERAL ANTI-MONEY LAUNDERING POLICIES, PRACTICES AND PROCEDURES
set_y(47) # Appointed Compliance Officer
safe_set(48, 1, 'Does your company AML policy cover CDD, EDD, UBO, Independent Testing, Periodic Review, Risk Assessment, Sanctions, PEP, Adverse Media, Transaction Monitoring, SARs, Training: Yes - Fully Implemented')
safe_set(49, 1, 'How many full time employees in Compliance Dept: 1 Full-Time Founder & Managing Director overseeing compliance functions, supported by external legal/compliance advisors')
safe_set(50, 1, 'Approved annually by Board: Yes - Approved by executive management / directors')
safe_set(51, 1, 'Board receives regular reporting: Yes - Regular management information provided to executive team')
safe_set(52, 1, 'Uses third parties for AML/KYC components: Yes - Identity verification, PEP/Sanctions screening, and infrastructure services through regulated partners (GTBank, Zenith Bank, Grey Business, Fincra, NOWPayments, Anchor)')
safe_set(53, 1, 'Automated API integrations with regulated banking & identity verification partners.')
set_y(54) # Representative of all branches
safe_set(55, 1, 'N/A')
safe_set(56, 1, 'AML/CTF framework is designed with reference to industry best practices, FATF guidelines, and applicable regulatory requirements.')

set_y(58) # Money laundering prevention
set_y(59) # Terrorist financing prevention
set_y(60) # Sanctions violations prevention
set_y(61) # Policies updated annually
set_y(64) # Gapped against US Standards
safe_set(65, 1, 'Yes - Retained in compliance repository')
set_y(66) # Gapped against EU Standards
safe_set(67, 1, 'Yes - Retained in compliance repository')

set_y(70) # Prohibit anonymous accounts
set_y(71) # Prohibit unlicensed banks
set_y(72) # Prohibit entities serving unlicensed banks
set_y(73) # Prohibit shell banks
set_y(74) # Prohibit entities serving shell banks
set_y(75) # Prohibit unlicensed remitters
set_y(76) # Assess PEP risks
set_y(77) # Define escalation processes
set_y(78) # Define termination processes
set_y(79) # Specify suspicious activity escalation
set_y(80) # Outline screening processes
set_y(81) # Maintain internal watchlists
set_y(82) # Define risk tolerance statement
set_y(83) # Record retention procedures (5+ years)
set_y(84) # Representative of all branches
safe_set(85, 1, 'N/A')
safe_set(86, 1, 'Record retention procedures comply with applicable laws for a minimum of 5 to 7 years in encrypted databases.')

# 5. AML, CTF & SANCTIONS RISK ASSESSMENT
set_y(89) # EWRA Client risk
set_y(90) # EWRA Product risk
set_y(91) # EWRA Channel risk
set_y(92) # EWRA Geography risk
set_y(95) # EWRA TM controls
set_y(96) # EWRA CDD controls
set_y(97) # EWRA PEP controls
set_y(98) # EWRA Transaction screening
set_y(99) # EWRA Adverse media screening
set_y(100) # EWRA Training
set_y(101) # EWRA Governance
set_y(102) # EWRA MI
set_y(103) # Completed in last 12 months (EWRA Q1 2026)
safe_set(104, 1, 'N/A - Enterprise-Wide Risk Assessment completed Q1 2026.')
set_y(106) # Representative of all branches
safe_set(107, 1, 'N/A')
safe_set(108, 1, 'Enterprise-Wide Risk Assessment reviewed annually.')

# 6. ANTI-BRIBERY AND CORRUPTION (ABC)
set_y(112) # ABC policies documented
set_y(113) # Enhanced public official requirements
set_y(114) # Designated ABC officer
set_y(117) # Prohibits bribes
set_y(118) # Prohibits falsification of records
set_y(119) # Controls to monitor ABC effectiveness
set_y(120) # Board receives regular ABC MI
set_y(121) # Enterprise ABC risk assessment
safe_set(122, 1, 'Every 12 Months (Annually)')
set_y(123) # ABC residual risk rating
set_y(126) # Intermediary liability
set_y(127) # Country/industry corruption risk
set_y(128) # State-owned entities / public officials
set_y(129) # Gifts/hospitality/donations/political
set_y(130) # Business activity changes
set_y(131) # Internal audit covers ABC
set_y(133) # Mandatory ABC training - Board
set_y(134) # Mandatory ABC training - Employees
set_y(135) # Mandatory ABC training - 3rd parties
set_y(136) # Mandatory ABC training - Contractors
set_y(137) # Targeted role ABC training
safe_set(138, 1, 'Reviewed upon any material operational change.')
set_y(139) # Representative of all branches
safe_set(140, 1, 'N/A')
safe_set(141, 1, 'Zero-tolerance policy against bribery and corruption enforced across all operations.')

# 7. COMPANY'S SOURCE OF FUNDS
safe_set(147, 1, 'Name of person/entity contributing funds: Aghogho Jossy Oboh (Founder & Managing Director)')
safe_set(148, 1, 'Source of funds contributed: Founder Equity Capital & Retained Business Working Capital')

# 8. DIRECTORS & AUTHORISED SIGNATORIES
# Director 1 (Aghogho Jossy Oboh)
safe_set(156, 2, 'Aghogho Jossy Oboh')                          # B156: Name
safe_set(156, 9, '15 March 1994')                               # I156: DOB
safe_set(157, 2, 'Founder & Managing Director')                # B157: Designation
safe_set(157, 4, 'Yes')                                         # D157: Authorised Signatory
safe_set(158, 2, '10 Winnie Okia Street, Effurun, Delta State, Nigeria') # B158: Address
safe_set(159, 3, 'Nigeria')                                     # C159: Country
safe_set(159, 7, '330102')                                      # G159: Post Code
safe_set(160, 2, 'Nigerian')                                    # B160: Nationality
safe_set(161, 3, 'Attached (Certified Copy of Passport / National ID & Proof of Address)') # C161
safe_set(162, 5, 'No')                                          # E162: PEP status

# Director 2 & 3
safe_set(164, 2, 'N/A (Single Director Entity)')
safe_set(172, 2, 'N/A (Single Director Entity)')

# 9. ULTIMATE BENEFICIAL OWNER (UBO)
# UBO 1 (Aghogho Jossy Oboh)
safe_set(184, 2, 'Aghogho Jossy Oboh')                          # B184: Name
safe_set(184, 9, '15 March 1994')                               # I184: DOB
safe_set(185, 2, 'Founder, Managing Director & 100% Shareholder') # B185: Designation
safe_set(185, 4, 'Yes')                                         # D185: Authorised Signatory
safe_set(186, 2, '10 Winnie Okia Street, Effurun, Delta State, Nigeria') # B186: Address
safe_set(187, 3, 'Nigeria')                                     # C187: Country
safe_set(187, 7, '330102')                                      # G187: Post Code
safe_set(188, 2, 'Nigerian')                                    # B188: Nationality
safe_set(189, 3, 'Attached (Certified Copy of Passport / National ID & Proof of Address)') # C189
safe_set(190, 5, 'No')                                          # E190: PEP status

safe_set(192, 2, 'N/A (No additional UBO >5%)')

# 10. KYC, CDD AND EDD
set_y(202) # Documented CDD policies
set_y(205) # CDD covers natural/legal persons
set_y(206) # CDD gathered prior to relationship
set_y(207) # Verified using reliable data
set_y(208) # Identify UBOs
set_y(209) # Verify UBOs
set_y(210) # Screening watchlists
set_y(211) # Screening PEPs
set_y(212) # Ongoing CDD
safe_set(214, 1, "Do the Entity's policies set out when CDD must be completed: Completed prior to relationship establishment / onboarding.")
set_y(216) # Ownership structure
set_y(217) # Customer identification
set_y(218) # Expected activity
set_y(219) # Nature of business/employment
set_y(220) # Product usage
set_y(221) # Purpose and nature of relationship
set_y(223) # UBO identified
set_y(224) # UBO verified
set_y(225) # Authorized signatories
set_y(226) # Key controllers
set_y(227) # Other relevant parties
safe_set(228, 1, 'What is minimum threshold applied to beneficial ownership: 10%')

set_y(229) # Risk classification
set_y(233) # Factors used
set_y(234) # Product usage factor
set_y(235) # Geography factor
set_y(236) # Business type factor
set_y(237) # Legal entity type factor
set_y(238) # Adverse info factor
safe_set(239, 1, 'Other factors: Transaction velocity & sanctions screening flags')

set_y(240) # Adverse media screening
set_y(243) # Onboarding
set_y(244) # KYC renewal
set_y(245) # Trigger event
safe_set(246, 1, 'Automated API screening through regulated provider databases.')

set_y(247) # PEP screening
set_y(250) # Onboarding
set_y(251) # KYC renewal
set_y(252) # Trigger event
safe_set(253, 1, 'Automated API screening against global PEP watchlists.')

set_y(254) # Review/update customer info
set_y(256) # KYC renewal
set_y(257) # Trigger event
set_y(258) # Metrics on reviews

set_y(260) # Non-account customer subject to EDD
set_y(261) # Non-resident customers subject to EDD
set_n(262) # Shell banks (Prohibited)
set_y(263) # PEPs subject to EDD
set_y(264) # PEP Related subject to EDD
set_y(265) # PEP Close Associate subject to EDD

set_y(267) # Non-account customer restricted
set_y(268) # Non-resident high-risk restricted
set_y(269) # Shell banks prohibited
set_y(270) # PEPs EDD required
set_y(271) # PEP Related EDD required
set_y(272) # PEP Close Associate EDD required

set_y(274) # EDD assessment elements
set_y(275) # Arms, defense (Restricted)
set_y(276) # Atomic power (Restricted)
set_y(277) # Extractive industries (Restricted)
set_y(278) # Precious metals (Restricted)
set_y(279) # Unregulated charities (Prohibited)
set_y(280) # Regulated charities (EDD)
set_n(281) # Adult entertainment (Prohibited)
set_y(282) # Virtual currencies (Controlled / Regulated partners)
set_n(283) # Marijuana (Prohibited)
set_y(284) # Embassies (Restricted/EDD)
set_n(285) # Gambling (Prohibited/Restricted)
set_y(286) # Payment Service Provider (EDD)
safe_set(288, 1, 'Details of restriction: Prohibited categories (Adult entertainment, Marijuana, Shell Banks, Unregulated Gambling) are auto-blocked at onboarding.')

set_y(289) # Control review on EDD clients
set_y(291) # Representative of all branches
safe_set(292, 1, 'N/A')
safe_set(293, 1, 'Tiered CDD/EDD strictly enforced across all user levels.')

# 11. TRANSACTION MONITORING AND REPORTING
set_y(295) # Subject to regulatory reporting
safe_set(297, 1, 'If Yes, How often: Promptly upon detection of suspicious activity; monthly summary CTR reports.')
set_y(298) # Risk-based TM processes
safe_set(300, 1, 'Automated real-time transaction monitoring engine combined with management compliance review.')
safe_set(301, 1, 'High-value fiat withdrawals (>USD 10,000 equivalent) and manual SAR escalation reviews.')
set_y(302) # STR regulatory requirements
set_y(304) # STR policies and procedures
set_y(305) # Review and escalate TM
set_y(306) # Representative of all branches
safe_set(307, 1, 'N/A')
safe_set(308, 1, 'Customer funds are maintained separately from company operating funds through regulated banking and payment partners in accordance with applicable agreements.')

# 12. DATA PROTECTION
set_y(310) # Information security/privacy controls designed with reference to best practices
safe_set(311, 1, 'Information security and privacy controls are designed with reference to industry best practices (NDPR / GDPR framework alignment).')
safe_set(312, 1, 'Data protection responsibilities are currently overseen by management. A dedicated Data Protection Officer will be appointed as the business scales and where required by applicable regulations.')
set_y(313) # Respond to data subject requests
set_y(314) # Incident management & breach register
set_y(315) # Conducted DPIA
set_y(316) # Enforce Data Processing Agreements
set_y(317) # Legal basis documented (Privacy Policy / ROPA)
set_y(318) # Privacy notice provided at collection
safe_set(319, 1, 'Attached (Data Protection & Privacy Policy)')

# 13. COMPLIANCE AND REGULATORY
set_y(323) # Compliance oversight function
set_y(324) # Management oversight
set_y(325) # List of legislative/regulatory requirements

# 14. INFORMATION SECURITY
set_y(328) # DR plan and secondary DR site
set_y(330) # DR plan & DR site
set_y(331) # Business continuity plans
set_y(332) # Incident response team & procedures
set_y(333) # Info sec training for staff
set_y(334) # Manage third party risk
set_y(335) # Documented records retention & backup policy
set_y(336) # Secure disposal of info
set_y(337) # ISMS in place
set_y(338) # Security measures documented & tested
safe_set(339, 1, 'Attached (Information Security Framework)')

# 15. SANCTIONS
set_y(342) # Sanctions policy approved
set_y(343) # Sanctions policy approved by management
set_y(344) # Prevent use of accounts causing violation
set_y(345) # Detect sanctions evasion
set_y(346) # Screen customers & UBOs
safe_set(347, 1, 'Automated real-time API screening against global watchlists.')
set_y(348) # Screen cross-border data
set_y(349) # Screen cross-border data
safe_set(350, 1, 'Automated real-time cross-border transaction screening engine.')
set_y(351) # Sanctions lists used
set_y(353) # UN List
set_y(354) # OFAC List
set_y(355) # OFSI HMT List
set_y(356) # EU List
set_y(357) # G7 Lists
safe_set(358, 1, 'UK HMT, FATF, and Interpol watchlists.')
safe_set(360, 1, 'Customer Data: Real-time automated API updates')
safe_set(361, 1, 'Transactions: Real-time automated API updates')
set_y(362) # Physical presence in sanctioned regions header
set_n(363) # Physical presence in sanctioned regions (NO)
set_y(364) # Representative of all branches
safe_set(365, 1, 'N/A')
safe_set(366, 1, 'Sanctioned jurisdictions strictly prohibited & geo-blocked.')

# 16. TRAINING AND EDUCATION
set_y(370) # Mandatory training header
set_y(371) # Reporting training
set_y(372) # ML/TF examples training
set_y(373) # Internal policies training
set_y(374) # Market issues training
set_y(375) # Conduct training
set_y(377) # Board training
set_y(378) # 1st line training
set_y(379) # 2nd line training
set_y(380) # 3rd line training
set_y(381) # Targeted role training
set_y(382) # Customized compliance training
set_y(383) # Representative of all branches
safe_set(384, 1, 'N/A')
safe_set(385, 1, 'Annual training required for operational personnel.')

# 17. QUALITY ASSURANCE / COMPLIANCE TESTING
set_y(388) # QA testing header
set_y(389) # KYC QA testing
set_y(390) # Compliance testing process
set_y(391) # Representative of all branches
safe_set(392, 1, 'N/A')
safe_set(393, 1, 'Internal compliance testing conducted periodically.')

# 18. INDEPENDENT AUDIT AND FINDINGS
set_n(397) # Examined by regulatory body (NO - Pre-launch startup)
safe_set(398, 1, 'Internal compliance review conducted periodically; external 3rd-party independent audit planned post-launch.')
set_y(399) # Internal audit function header
set_y(400) # Internal audit assessing FCC/AML
safe_set(402, 1, 'Internal Audit: Quarterly')
safe_set(403, 1, 'External Third Party: Annually (Planned post-launch)')
set_y(405) # Audit covers areas header
set_y(406) # Policies
set_y(407) # KYC/CDD/EDD
set_y(408) # TM
set_y(409) # Screening
set_y(410) # Name screening
set_y(411) # Training
set_y(412) # Tech
set_y(413) # Governance
set_y(414) # Reporting
set_y(415) # SAR filing
set_y(416) # EWRA
safe_set(417, 1, 'Data privacy & treasury reconciliation.')
set_y(418) # Adverse findings tracked header
set_y(419) # Adverse findings tracked
set_y(420) # Representative of all branches
safe_set(421, 1, 'N/A')
safe_set(422, 1, 'Audit findings reported directly to executive management.')

# 19. DECLARATION AND CONSENT
safe_set(427, 2, 'Aghogho Jossy Oboh')                          # B427: Full Name
safe_set(427, 7, 'Aghogho Jossy Oboh')                          # G427: Full Name (Right side)
safe_set(428, 2, 'Aghogho Jossy Oboh (Digitally Signed)')        # B428: Signature
safe_set(428, 7, 'Aghogho Jossy Oboh (Digitally Signed)')        # G428: Signature (Right side)
safe_set(429, 2, 'Founder & Managing Director')                # B429: Role
safe_set(429, 7, 'Founder & Managing Director')                # G429: Role (Right side)
safe_set(430, 2, '7 August 2026')                              # B430: Date
safe_set(430, 7, '7 August 2026')                              # G430: Date (Right side)

# ---------------------------------------------------------
# SHEET 2: List of documents to be attache
# ---------------------------------------------------------
ws2 = wb['List of documents to be attache']
for row_idx in range(3, 15):
    ws2.cell(row_idx, 2).value = 'Attached'

# Save both output files
wb.save(out_path1)
wb.save(out_path2)

print('=== POPULATION COMPLETED SUCCESSFULLY (CORRECTED & ACCURATE) ===')
print('Output 1:', out_path1)
print('Output 2:', out_path2)
